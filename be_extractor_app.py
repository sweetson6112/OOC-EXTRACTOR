"""
Bill of Entry PDF Extractor – Streamlit App
Extracts key data from Indian Customs Bill of Entry (Warehouse BE) PDFs.
Run: streamlit run be_extractor_app.py
"""

import re, io
import streamlit as st
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="BE PDF Extractor", page_icon="🛃", layout="wide")
st.title("🛃 Indian Customs – Bill of Entry PDF Extractor")
st.caption("Upload one or more Warehouse BE (Out-of-Charge) PDFs to extract header, invoice & item data into Excel.")

# ── helpers ──────────────────────────────────────────────────────────────────
def first(text, pattern, default=""):
    m = re.search(pattern, text, re.IGNORECASE)
    return m.group(1).strip() if m else default

def extract_be_data(pdf_bytes: bytes) -> dict:
    """Extract all relevant fields from a BE PDF."""
    all_text = ""
    pages_text = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            pages_text.append(t)
            all_text += t + "\n"

    data = {}

    # ── HEADER ────────────────────────────────────────────────────────────
    data["BE No"]              = first(all_text, r"BE No\s+BE Date.*?\n.*?(\d{7})")
    data["BE Date"]            = first(all_text, r"(\d{2}/\d{2}/\d{4})\n")
    data["BE Type"]            = first(all_text, r"BE Type\s*\n\s*([A-Z])")
    data["Port Code"]          = first(all_text, r"Port Code\s*\n\s*(\w+)")
    data["IEC/Br"]             = first(all_text, r"IEC/Br\s*\n\s*([\w/]+)")
    data["GSTIN"]              = first(all_text, r"GSTIN/TYPE\s*\n\s*(\w+)")
    data["CB Code"]            = first(all_text, r"CB CODE\s*\n\s*(\w+)")
    data["Importer"]           = first(all_text, r"1\.IMPORTER\s+NAME\s*&\s*ADDRESS\s*\n([^\n]+)")
    data["CB Name"]            = first(all_text, r"2\.CB NAME\s+([^\n]+)")
    data["Country of Origin"]  = first(all_text, r"13\.COUNTRY OF ORIGIN\s+([^\n]+14\.)")
    if data["Country of Origin"]: data["Country of Origin"] = data["Country of Origin"].replace("14.","").strip()
    data["Country of Consignment"] = first(all_text, r"14\.COUNTRY OF CONSIGNMENT\s*\n([^\n]+)")
    data["Port of Loading"]    = first(all_text, r"15\.PORT OF LOADING\s+([^\n]+16\.)")
    if data["Port of Loading"]: data["Port of Loading"] = data["Port of Loading"].replace("16.","").strip()
    data["Port of Shipment"]   = first(all_text, r"16\.PORT OF SHIPMENT\s*\n([^\n]+)")
    data["IGM No"]             = first(all_text, r"(\d{7})\s+\d{2}/\d{2}/\d{4}\s+\d{2}/\d{2}/\d{4}")
    data["IGM Date"]           = first(all_text, r"\d{7}\s+(\d{2}/\d{2}/\d{4})\s+\d{2}/\d{2}/\d{4}")
    data["INW Date"]           = first(all_text, r"\d{7}\s+\d{2}/\d{2}/\d{4}\s+(\d{2}/\d{2}/\d{4})")
    data["MAWB No"]            = first(all_text, r"6\.MAWB NO\s*\n?([\w]+)")
    data["MAWB Date"]          = first(all_text, r"7\.DATE\s*\n?(\d{2}/\d{2}/\d{4})")
    data["GW (KGS)"]           = first(all_text, r"G\.WT\s*\(KGS\)\s*\n?(\d+)")
    data["Packages"]           = first(all_text, r"PKG\s*\n?(\d+)")
    data["Exchange Rate"]      = first(all_text, r"(1 USD=[\d\.]+INR)")
    data["OOC No"]             = first(all_text, r"OOC NO\.\s*\n?([\d]+)")
    data["OOC Date"]           = first(all_text, r"OOC DATE\s*\n?(\d{2}-\d{2}-\d{4})")

    # Duty summary
    duty_block = re.search(r"([\d\.]+)\s+([\d\.]+)\s+[\d\.]+\s+[\d\.]+\s+([\d\.]+)\s+[\d\.]+\s+([\d\.]+)", all_text)
    data["BCD (INR)"]    = first(all_text, r"1\.BCD.*?\n\s*([\d,\.]+)")
    data["SWS (INR)"]    = first(all_text, r"3\.SWS.*?\n\s*([\d,\.]+)")
    data["IGST (INR)"]   = first(all_text, r"7\.IGST.*?\n\s*([\d,\.]+)")
    data["TOT ASS VAL"]  = first(all_text, r"18\.TOT\.ASS VAL\s*\n?([\d,\.]+)")
    data["TOT AMOUNT"]   = first(all_text, r"19\.TOT\. AMOUNT\s*\n?([\d,\.]+)")
    data["FINE"]         = first(all_text, r"17\.FINE\s*\n?([\d,\.]+)")

    # Container
    data["Container No"] = first(all_text, r"5\.CONTAINER NUMBER\s*\n?\s*([A-Z]{4}\d+)")
    data["Seal No"]      = first(all_text, r"4\.SEAL\s*\n?\s*(\d+)")

    # ── INVOICES ──────────────────────────────────────────────────────────
    # Pattern: look for invoice number + amount lines
    inv_pattern = re.findall(
        r"(\d)\s+(13200\d+)\s+([\d,\.]+)\s+(USD|EUR|GBP)",
        all_text)
    invoices = []
    for sno, invno, amt, cur in inv_pattern:
        invoices.append({
            "S.No": int(sno),
            "Invoice No": invno,
            "Invoice Amount": float(amt.replace(",","")),
            "Currency": cur,
        })
    data["invoices"] = invoices

    # ── ITEMS ─────────────────────────────────────────────────────────────
    # Parse Part-III pages which contain assess value & total duty per item
    items = []
    # Find all item blocks: look for "InvSNo ItemSNo CTH" patterns
    item_blocks = re.findall(
        r"(\d)\s+(\d+)\s+(1\d{7})\s+NOEXCISE\s+([^\n]+)\n"
        r".*?([\d\.]+)\s+([A-Z]{2})\s+([\d,\.]+)\s+KGS.*?\n"
        r".*?29\.ASSESS VALUE\s*\n\s*([\d,\.]+)\s+([\d,\.]+)",
        all_text, re.DOTALL)
    for m in item_blocks:
        invsno, itemsno, cth, desc, upi, coo, qty, av, td = m
        items.append({
            "Inv S.No": int(invsno),
            "Item S.No": int(itemsno),
            "CTH": cth,
            "Description": desc.strip()[:120],
            "UPI": float(upi),
            "COO": coo,
            "Qty (KGS)": float(qty.replace(",","")),
            "Assess Value (INR)": float(av.replace(",","")),
            "Total Duty (INR)": float(td.replace(",","")),
        })
    data["items"] = items

    return data


def to_excel(all_data: list[dict]) -> bytes:
    wb = Workbook()
    HDR  = PatternFill("solid", start_color="1F4E79")
    HDR2 = PatternFill("solid", start_color="2E75B6")
    ALT  = PatternFill("solid", start_color="DEEAF1")
    SUM  = PatternFill("solid", start_color="FFF2CC")
    OK   = PatternFill("solid", start_color="E2EFDA")
    ERR  = PatternFill("solid", start_color="FFE0E0")
    BD   = Border(left=Side(style="thin"),right=Side(style="thin"),
                  top=Side(style="thin"),bottom=Side(style="thin"))

    def hcell(ws, r, c, v, fill=None, span=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = fill or HDR
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BD
        return cell

    def dcell(ws, r, c, v, bold=False, fill=None, fmt=None, align="left"):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(bold=bold, size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = BD
        if fill: cell.fill = fill
        if fmt:  cell.number_format = fmt
        return cell

    # ── Sheet: All BE Headers ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "BE Headers"
    header_fields = ["BE No","BE Date","BE Type","Port Code","IEC/Br","GSTIN",
                     "CB Code","Importer","CB Name","Country of Origin",
                     "Country of Consignment","Port of Loading","Port of Shipment",
                     "IGM No","IGM Date","INW Date","MAWB No","MAWB Date",
                     "GW (KGS)","Packages","Exchange Rate",
                     "BCD (INR)","SWS (INR)","IGST (INR)",
                     "TOT ASS VAL","TOT AMOUNT","FINE",
                     "Container No","Seal No","OOC No","OOC Date"]
    ws.column_dimensions["A"].width = 28
    hcell(ws, 1, 1, "Field")
    for i, d in enumerate(all_data, 2):
        ws.column_dimensions[get_column_letter(i)].width = 25
        hcell(ws, 1, i, f"BE {d.get('BE No','#'+str(i-1))}")
    for ri, f in enumerate(header_fields, 2):
        fill = ALT if ri % 2 == 0 else None
        dcell(ws, ri, 1, f, bold=True, fill=fill)
        for ci, d in enumerate(all_data, 2):
            dcell(ws, ri, ci, d.get(f,""), fill=fill,
                  align="right" if f in ("GW (KGS)","Packages") else "left")

    # ── Sheet: All Invoices ───────────────────────────────────────────────
    ws2 = wb.create_sheet("All Invoices")
    inv_cols = ["BE No","S.No","Invoice No","Invoice Amount","Currency"]
    inv_w    = [15,6,18,18,10]
    for i,(c,w) in enumerate(zip(inv_cols,inv_w),1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        hcell(ws2, 1, i, c)
    row = 2
    for d in all_data:
        for inv in d.get("invoices",[]):
            fill = ALT if row % 2 == 0 else None
            dcell(ws2, row, 1, d.get("BE No",""), fill=fill, align="center")
            dcell(ws2, row, 2, inv["S.No"],       fill=fill, align="center")
            dcell(ws2, row, 3, inv["Invoice No"], fill=fill)
            dcell(ws2, row, 4, inv["Invoice Amount"], fill=fill, fmt="#,##0.00", align="right")
            dcell(ws2, row, 5, inv["Currency"],   fill=fill, align="center")
            row += 1
    # Total
    dcell(ws2, row, 1, "TOTAL", bold=True, fill=SUM, align="center")
    ws2.merge_cells(f"A{row}:C{row}")
    dcell(ws2, row, 4, f"=SUM(D2:D{row-1})", bold=True, fill=SUM, fmt="#,##0.00", align="right")
    dcell(ws2, row, 5, "", fill=SUM)

    # ── Sheet: All Items ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("All Items")
    it_cols = ["BE No","Inv S.No","Item S.No","CTH","Description","Qty (KGS)",
               "UPI (USD)","COO","Assess Value (INR)","Total Duty (INR)"]
    it_w    = [14,9,9,12,55,10,12,6,20,18]
    for i,(c,w) in enumerate(zip(it_cols,it_w),1):
        ws3.column_dimensions[get_column_letter(i)].width = w
        hcell(ws3, 1, i, c, fill=HDR2)
    row = 2
    for d in all_data:
        for it in d.get("items",[]):
            fill = ALT if row % 2 == 0 else None
            dcell(ws3, row, 1, d.get("BE No",""),  fill=fill, align="center")
            dcell(ws3, row, 2, it["Inv S.No"],     fill=fill, align="center")
            dcell(ws3, row, 3, it["Item S.No"],    fill=fill, align="center")
            dcell(ws3, row, 4, it["CTH"],          fill=fill, align="center")
            dcell(ws3, row, 5, it["Description"],  fill=fill)
            dcell(ws3, row, 6, it["Qty (KGS)"],    fill=fill, fmt="#,##0.000", align="right")
            dcell(ws3, row, 7, it["UPI"],          fill=fill, fmt="#,##0.000000", align="right")
            dcell(ws3, row, 8, it["COO"],          fill=fill, align="center")
            dcell(ws3, row, 9, it["Assess Value (INR)"],fill=fill, fmt="#,##0.00", align="right")
            dcell(ws3, row,10, it["Total Duty (INR)"],  fill=fill, fmt="#,##0.00", align="right")
            row += 1
    dcell(ws3, row, 1, "TOTAL", bold=True, fill=SUM, align="center")
    ws3.merge_cells(f"A{row}:H{row}")
    dcell(ws3, row, 9, f"=SUM(I2:I{row-1})", bold=True, fill=SUM, fmt="#,##0.00", align="right")
    dcell(ws3, row,10, f"=SUM(J2:J{row-1})", bold=True, fill=SUM, fmt="#,##0.00", align="right")

    # ── Sheet: Duty Reconciliation ─────────────────────────────────────────
    ws4 = wb.create_sheet("Duty Reconciliation")
    rec_cols = ["BE No","Sum Assess Value (INR)","TOT ASS VAL (Header)",
                "Diff – AV","Sum Total Duty (INR)","TOT AMOUNT (Header)",
                "Diff – Duty","FINE (Header)","Status"]
    rec_w   = [15,22,22,16,22,22,16,16,14]
    for i,(c,w) in enumerate(zip(rec_cols,rec_w),1):
        ws4.column_dimensions[get_column_letter(i)].width = w
        hcell(ws4, 1, i, c)

    for ri, d in enumerate(all_data, 2):
        items    = d.get("items",[])
        sum_av   = sum(x["Assess Value (INR)"] for x in items)
        sum_duty = sum(x["Total Duty (INR)"] for x in items)
        try:    hdr_av   = float(str(d.get("TOT ASS VAL","0")).replace(",",""))
        except: hdr_av   = 0
        try:    hdr_duty = float(str(d.get("TOT AMOUNT","0")).replace(",",""))
        except: hdr_duty = 0
        try:    fine     = float(str(d.get("FINE","0")).replace(",",""))
        except: fine     = 0
        diff_av   = round(sum_av   - hdr_av,   2)
        diff_duty = round(sum_duty - hdr_duty, 2)
        ok = abs(diff_av) < 1 and abs(diff_duty) < 1
        fill = OK if ok else ERR
        vals = [d.get("BE No",""), sum_av, hdr_av, diff_av,
                sum_duty, hdr_duty, diff_duty, fine,
                "✔ MATCH" if ok else "⚠ MISMATCH"]
        for ci, v in enumerate(vals, 1):
            fmt = "#,##0.00" if ci in (2,3,4,5,6,7,8) else None
            al  = "right"   if ci in (2,3,4,5,6,7,8) else "center"
            dcell(ws4, ri, ci, v, fill=fill, fmt=fmt, align=al,
                  bold=(ci==9))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit UI ──────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "Upload Bill of Entry PDF(s)",
    type="pdf",
    accept_multiple_files=True,
    help="Supports Indian Customs Warehouse BE PDFs (any number of invoices)"
)

if uploaded:
    all_data = []
    progress = st.progress(0)
    for i, f in enumerate(uploaded):
        with st.spinner(f"Processing {f.name} …"):
            try:
                d = extract_be_data(f.read())
                d["_filename"] = f.name
                all_data.append(d)
                st.success(f"✅ {f.name} — BE No: {d.get('BE No','?')} | {len(d.get('invoices',[]))} invoices | {len(d.get('items',[]))} items")
            except Exception as e:
                st.error(f"❌ {f.name}: {e}")
        progress.progress((i+1)/len(uploaded))

    if all_data:
        st.divider()
        tab1, tab2, tab3, tab4 = st.tabs(["📋 BE Header","🧾 Invoices","📦 Items","⚖️ Duty Reconciliation"])

        with tab1:
            hdr_rows = []
            for d in all_data:
                hdr_rows.append({
                    "BE No": d.get("BE No"), "BE Date": d.get("BE Date"),
                    "Importer": d.get("Importer"), "IGM No": d.get("IGM No"),
                    "IGM Date": d.get("IGM Date"), "INW Date": d.get("INW Date"),
                    "MAWB No": d.get("MAWB No"), "GW KGS": d.get("GW (KGS)"),
                    "Container": d.get("Container No"), "Seal": d.get("Seal No"),
                    "TOT ASS VAL": d.get("TOT ASS VAL"), "TOT AMOUNT": d.get("TOT AMOUNT"),
                    "OOC No": d.get("OOC No"), "OOC Date": d.get("OOC Date"),
                })
            st.dataframe(pd.DataFrame(hdr_rows), use_container_width=True)

        with tab2:
            inv_rows = []
            for d in all_data:
                for inv in d.get("invoices",[]):
                    inv_rows.append({"BE No": d.get("BE No"), **inv})
            st.dataframe(pd.DataFrame(inv_rows), use_container_width=True)

        with tab3:
            item_rows = []
            for d in all_data:
                for it in d.get("items",[]):
                    item_rows.append({"BE No": d.get("BE No"), **it})
            st.dataframe(pd.DataFrame(item_rows), use_container_width=True)

        with tab4:
            rec_rows = []
            for d in all_data:
                items    = d.get("items",[])
                sum_av   = round(sum(x["Assess Value (INR)"] for x in items),2)
                sum_duty = round(sum(x["Total Duty (INR)"]   for x in items),2)
                try:    hdr_av   = float(str(d.get("TOT ASS VAL","0")).replace(",",""))
                except: hdr_av   = 0
                try:    hdr_duty = float(str(d.get("TOT AMOUNT","0")).replace(",",""))
                except: hdr_duty = 0
                try:    fine     = float(str(d.get("FINE","0")).replace(",",""))
                except: fine     = 0
                diff_av   = round(sum_av   - hdr_av,   2)
                diff_duty = round(sum_duty - hdr_duty, 2)
                ok = abs(diff_av)<1 and abs(diff_duty)<1
                rec_rows.append({
                    "BE No": d.get("BE No"),
                    "Sum AV (items)": sum_av, "TOT ASS VAL (hdr)": hdr_av, "Diff AV": diff_av,
                    "Sum Duty (items)": sum_duty, "TOT AMOUNT (hdr)": hdr_duty, "Diff Duty": diff_duty,
                    "FINE": fine, "Status": "✔ MATCH" if ok else "⚠ MISMATCH"
                })
            df_rec = pd.DataFrame(rec_rows)
            st.dataframe(
                df_rec.style.apply(
                    lambda r: ["background-color:#E2EFDA"]*len(r) if r["Status"]=="✔ MATCH"
                              else ["background-color:#FFE0E0"]*len(r), axis=1),
                use_container_width=True)

        st.divider()
        excel_bytes = to_excel(all_data)
        st.download_button(
            label="⬇️ Download Excel Report",
            data=excel_bytes,
            file_name="BE_Extract_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary")
else:
    st.info("👆 Upload one or more Bill of Entry PDFs to get started.")
    with st.expander("ℹ️ How to use"):
        st.markdown("""
**Extracted data:**
- **BE Header**: BE No, Date, IGM, MAWB, Container, Seal, Duty Summary, OOC details
- **Invoices**: All invoice numbers, amounts, currencies
- **Items**: Each line item with CTH, description, quantity, unit price, COO, assess value & total duty
- **Duty Reconciliation**: Compares sum of item-level assess values & duties against BE header totals (18.TOT.ASS VAL & 19.TOT.AMOUNT). Flags any mismatch; if diff ≠ 0 it should equal 17.FINE.

**Requirements:**
```
pip install streamlit pdfplumber openpyxl pandas
```
Run: `streamlit run be_extractor_app.py`
        """)
