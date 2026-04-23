"""
Indian Customs – Bill of Entry PDF Extractor
=============================================
Streamlit app that extracts data from Warehouse BE (Out-of-Charge) PDFs
issued by Indian Customs via ICEGATE.

Run locally
-----------
    pip install -r requirements.txt
    streamlit run app.py

Deploy to Streamlit Community Cloud
------------------------------------
1. Push this repo to GitHub (public or private).
2. Go to https://share.streamlit.io  →  New app
   → select repo  →  main file: app.py
3. Click Deploy.
"""

import re
import io
import streamlit as st
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="BE PDF Extractor – Indian Customs",
    page_icon="🛃",
    layout="wide",
)


# ─────────────────────────────────────────────────────────────────────────────
# Extraction helpers
# ─────────────────────────────────────────────────────────────────────────────

def _first(text: str, pattern: str, default: str = "") -> str:
    m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    return m.group(1).strip() if m else default


def _safe_float(val: str) -> float:
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def extract_be_data(pdf_bytes: bytes) -> dict:
    """
    Parse a single Warehouse BE PDF and return a structured dict.
    Works with any number of invoices / items.
    """
    full_text = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

    d: dict = {}

    # ── Header fields ─────────────────────────────────────────────────────
    d["BE No"]    = _first(full_text, r"BE No\s+BE Date[^\n]*\n\s*(\d{7,8})")
    d["BE Date"]  = _first(full_text, r"BE No\s+BE Date[^\n]*\n\s*\d+\s+(\d{2}/\d{2}/\d{4})")
    d["BE Type"]  = _first(full_text, r"BE Type\s*\n\s*([A-Z])\b")
    d["Port Code"]= _first(full_text, r"Port Code\s*\n\s*(\w+)")
    d["IEC/Br"]   = _first(full_text, r"IEC/Br\s*\n\s*([\w/]+)")
    d["GSTIN"]    = _first(full_text, r"GSTIN/TYPE\s*\n\s*([\w/]+)")
    d["CB Code"]  = _first(full_text, r"CB CODE\s*\n\s*([\w]+)")
    d["Importer"] = _first(full_text, r"1\.IMPORTER\s+NAME\s*&\s*ADDRESS\s*\n([\w &,\.\-/]+)")
    d["CB Name"]  = _first(full_text, r"2\.CB NAME\s+([\w &]+)")

    coo_raw = _first(full_text, r"13\.COUNTRY OF ORIGIN\s+(.*?)14\.")
    d["Country of Origin"]      = re.sub(r"\s+", " ", coo_raw).strip()
    d["Country of Consignment"] = _first(full_text, r"14\.COUNTRY OF CONSIGNMENT\s*\n([\w ]+)")
    pol_raw = _first(full_text, r"15\.PORT OF LOADING\s+(.*?)16\.")
    d["Port of Loading"]        = re.sub(r"\s+", " ", pol_raw).strip()
    d["Port of Shipment"]       = _first(full_text, r"16\.PORT OF SHIPMENT\s*\n([\w ]+)")

    igm_m = re.search(r"(\d{7,10})\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})", full_text)
    if igm_m:
        d["IGM No"]   = igm_m.group(1)
        d["IGM Date"] = igm_m.group(2)
        d["INW Date"] = igm_m.group(3)
    else:
        d["IGM No"] = d["IGM Date"] = d["INW Date"] = ""

    d["MAWB No"]       = _first(full_text, r"6\.MAWB NO\s*\n?([\w]+)")
    d["MAWB Date"]     = _first(full_text, r"7\.DATE\s*\n?(\d{2}/\d{2}/\d{4})")
    d["GW (KGS)"]      = _first(full_text, r"G\.WT\s*\(KGS\)\s*\n?(\d+)")
    d["Packages"]      = _first(full_text, r"\bPKG\b\s*\n?(\d+)")
    d["Exchange Rate"] = _first(full_text, r"(1 USD=[\d\.]+INR)")

    # Duty summary numbers
    duty_m = re.search(
        r"([\d,\.]+)\s+([\d,\.]+)\s+([\d,\.]+)\s+[\d,\.]+\s+[\d,\.]+\s+[\d,\.]+\s+"
        r"([\d,\.]+)\s+[\d,\.]+\s+([\d,\.]+)",
        full_text,
    )
    if duty_m:
        d["BCD (INR)"]   = duty_m.group(1).replace(",", "")
        d["ACD (INR)"]   = duty_m.group(2).replace(",", "")
        d["SWS (INR)"]   = duty_m.group(3).replace(",", "")
        d["IGST (INR)"]  = duty_m.group(4).replace(",", "")
        d["TOT ASS VAL"] = duty_m.group(5).replace(",", "")
    else:
        d["BCD (INR)"] = d["ACD (INR)"] = d["SWS (INR)"] = ""
        d["IGST (INR)"] = d["TOT ASS VAL"] = ""

    d["TOT AMOUNT"] = _first(full_text, r"19\.TOT\.\s*AMOUNT\s*\n?([\d,\.]+)")
    d["FINE"]       = _first(full_text, r"17\.FINE\s*\n?([\d,\.]+)")
    d["Container No"]= _first(full_text, r"5\.CONTAINER NUMBER\s*\n?\s*([A-Z]{4}\d{7})")
    d["Seal No"]     = _first(full_text, r"4\.SEAL\s*\n?\s*(\d{6,})")
    d["OOC No"]      = _first(full_text, r"OOC NO\.?\s*\n?\s*(\d{9,12})")
    d["OOC Date"]    = _first(full_text, r"OOC DATE\s*\n?\s*(\d{2}-\d{2}-\d{4})")

    # ── Invoices ──────────────────────────────────────────────────────────
    inv_hits = re.findall(
        r"(\d)\s+(13\d{8})\s+([\d,\.]+)\s+(USD|EUR|GBP|INR)",
        full_text,
    )
    invoices = []
    seen_inv: set = set()
    for sno, invno, amt, cur in inv_hits:
        key = (invno, amt)
        if key not in seen_inv:
            seen_inv.add(key)
            invoices.append({
                "S.No": int(sno),
                "Invoice No": invno,
                "Invoice Amount (USD)": _safe_float(amt),
                "Currency": cur,
            })
    d["invoices"] = sorted(invoices, key=lambda x: x["S.No"])

    # ── Items (from Part-III duty pages) ──────────────────────────────────
    items = []
    chunks = re.split(r"(?=\d+\s+\d+\s+1\d{7}\s+NOEXCISE)", full_text)
    for chunk in chunks[1:]:
        m_head = re.match(
            r"(\d+)\s+(\d+)\s+(1\d{7})\s+NOEXCISE\s+(.+?)(?=\d+\.\d+|\n\d+\.\d+|\n[A-Z]{2}\b)",
            chunk, re.DOTALL,
        )
        if not m_head:
            continue
        invsno  = int(m_head.group(1))
        itemsno = int(m_head.group(2))
        cth     = m_head.group(3)
        desc    = re.sub(r"\s+", " ", m_head.group(4)).strip()[:150]

        upi_m = re.search(r"(\d+\.\d{2,6})\s+[A-Z]{2}\b", chunk)
        upi   = _safe_float(upi_m.group(1)) if upi_m else 0.0

        coo_m = re.search(r"\d+\.\d+\s+([A-Z]{2})\s+[\d,\.]+\s+KGS", chunk)
        coo   = coo_m.group(1) if coo_m else ""

        qty_m = re.search(r"([A-Z]{2})\s+([\d,\.]+)\s+KGS", chunk)
        qty   = _safe_float(qty_m.group(2)) if qty_m else 0.0

        av_m  = re.search(r"29\.ASSESS VALUE\s*\n?\s*([\d,\.]+)", chunk)
        av    = _safe_float(av_m.group(1)) if av_m else 0.0

        td_m  = re.search(r"30\.\s*TOTAL DUTY\s*\n?\s*([\d,\.]+)", chunk)
        td    = _safe_float(td_m.group(1)) if td_m else 0.0

        if av > 0 or td > 0:
            items.append({
                "Inv S.No": invsno,
                "Item S.No": itemsno,
                "CTH": cth,
                "Description": desc,
                "UPI (USD)": upi,
                "COO": coo,
                "Qty (KGS)": qty,
                "Assess Value (INR)": av,
                "Total Duty (INR)": td,
            })

    d["items"] = sorted(items, key=lambda x: (x["Inv S.No"], x["Item S.No"]))
    return d


# ─────────────────────────────────────────────────────────────────────────────
# Excel builder
# ─────────────────────────────────────────────────────────────────────────────

def _make_fills():
    return {
        "HDR":  PatternFill("solid", start_color="1F4E79"),
        "HDR2": PatternFill("solid", start_color="2E75B6"),
        "ALT":  PatternFill("solid", start_color="DEEAF1"),
        "SUM":  PatternFill("solid", start_color="FFF2CC"),
        "OK":   PatternFill("solid", start_color="E2EFDA"),
        "ERR":  PatternFill("solid", start_color="FFE0E0"),
    }


def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _hcell(ws, r, c, v, fill, bd, font_sz=10):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(bold=True, color="FFFFFF", size=font_sz)
    cell.fill      = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = bd
    return cell


def _dcell(ws, r, c, v, bd, bold=False, fill=None, fmt=None, align="left"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = bd
    if fill: cell.fill          = fill
    if fmt:  cell.number_format = fmt
    return cell


def build_excel(all_data: list) -> bytes:
    F  = _make_fills()
    BD = _border()
    wb = Workbook()

    # ── Sheet 1: BE Headers ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "BE Headers"
    header_fields = [
        "BE No", "BE Date", "BE Type", "Port Code", "IEC/Br", "GSTIN",
        "CB Code", "Importer", "CB Name", "Country of Origin",
        "Country of Consignment", "Port of Loading", "Port of Shipment",
        "IGM No", "IGM Date", "INW Date", "MAWB No", "MAWB Date",
        "GW (KGS)", "Packages", "Exchange Rate",
        "BCD (INR)", "ACD (INR)", "SWS (INR)", "IGST (INR)",
        "TOT ASS VAL", "TOT AMOUNT", "FINE",
        "Container No", "Seal No", "OOC No", "OOC Date",
    ]
    ws.column_dimensions["A"].width = 28
    _hcell(ws, 1, 1, "Field", F["HDR"], BD)
    for ci, d in enumerate(all_data, 2):
        ws.column_dimensions[get_column_letter(ci)].width = 28
        _hcell(ws, 1, ci, f"BE {d.get('BE No', '#' + str(ci - 1))}", F["HDR"], BD)
    for ri, f in enumerate(header_fields, 2):
        fill = F["ALT"] if ri % 2 == 0 else None
        _dcell(ws, ri, 1, f, BD, bold=True, fill=fill)
        for ci, d in enumerate(all_data, 2):
            _dcell(ws, ri, ci, d.get(f, ""), BD, fill=fill)

    # ── Sheet 2: All Invoices ─────────────────────────────────────────────
    ws2 = wb.create_sheet("All Invoices")
    for i, (col, w) in enumerate(zip(
        ["BE No", "S.No", "Invoice No", "Invoice Amount (USD)", "Currency"],
        [16, 6, 18, 22, 10]
    ), 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        _hcell(ws2, 1, i, col, F["HDR"], BD)
    row = 2
    for d in all_data:
        for inv in d.get("invoices", []):
            fill = F["ALT"] if row % 2 == 0 else None
            _dcell(ws2, row, 1, d.get("BE No", ""),        BD, fill=fill, align="center")
            _dcell(ws2, row, 2, inv["S.No"],                BD, fill=fill, align="center")
            _dcell(ws2, row, 3, inv["Invoice No"],          BD, fill=fill)
            _dcell(ws2, row, 4, inv["Invoice Amount (USD)"],BD, fill=fill, fmt="#,##0.00", align="right")
            _dcell(ws2, row, 5, inv["Currency"],            BD, fill=fill, align="center")
            row += 1
    _dcell(ws2, row, 1, "TOTAL", BD, bold=True, fill=F["SUM"], align="center")
    ws2.merge_cells(f"A{row}:C{row}")
    _dcell(ws2, row, 4, f"=SUM(D2:D{row - 1})", BD, bold=True, fill=F["SUM"], fmt="#,##0.00", align="right")
    _dcell(ws2, row, 5, "",                       BD, fill=F["SUM"])

    # ── Sheet 3: All Items ────────────────────────────────────────────────
    ws3 = wb.create_sheet("All Items")
    it_cols = ["BE No", "Inv S.No", "Item S.No", "CTH", "Description",
               "Qty (KGS)", "UPI (USD)", "COO", "Assess Value (INR)", "Total Duty (INR)"]
    for i, (col, w) in enumerate(zip(it_cols, [14, 9, 9, 12, 55, 10, 14, 6, 20, 18]), 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
        _hcell(ws3, 1, i, col, F["HDR2"], BD)
    row = 2
    for d in all_data:
        for it in d.get("items", []):
            fill = F["ALT"] if row % 2 == 0 else None
            _dcell(ws3, row, 1, d.get("BE No", ""),       BD, fill=fill, align="center")
            _dcell(ws3, row, 2, it["Inv S.No"],            BD, fill=fill, align="center")
            _dcell(ws3, row, 3, it["Item S.No"],           BD, fill=fill, align="center")
            _dcell(ws3, row, 4, it["CTH"],                 BD, fill=fill, align="center")
            _dcell(ws3, row, 5, it["Description"],         BD, fill=fill)
            _dcell(ws3, row, 6, it["Qty (KGS)"],           BD, fill=fill, fmt="#,##0.000", align="right")
            _dcell(ws3, row, 7, it["UPI (USD)"],           BD, fill=fill, fmt="#,##0.000000", align="right")
            _dcell(ws3, row, 8, it["COO"],                 BD, fill=fill, align="center")
            _dcell(ws3, row, 9, it["Assess Value (INR)"],  BD, fill=fill, fmt="#,##0.00", align="right")
            _dcell(ws3, row,10, it["Total Duty (INR)"],    BD, fill=fill, fmt="#,##0.00", align="right")
            row += 1
    _dcell(ws3, row, 1, "TOTAL", BD, bold=True, fill=F["SUM"], align="center")
    ws3.merge_cells(f"A{row}:H{row}")
    _dcell(ws3, row, 9, f"=SUM(I2:I{row - 1})", BD, bold=True, fill=F["SUM"], fmt="#,##0.00", align="right")
    _dcell(ws3, row,10, f"=SUM(J2:J{row - 1})", BD, bold=True, fill=F["SUM"], fmt="#,##0.00", align="right")

    # ── Sheet 4: Duty Reconciliation ──────────────────────────────────────
    ws4 = wb.create_sheet("Duty Reconciliation")
    rec_cols = ["BE No",
                "Sum Assess Value (INR)", "TOT ASS VAL (Header)", "Diff – AV",
                "Sum Total Duty (INR)",   "TOT AMOUNT (Header)",  "Diff – Duty",
                "FINE (Header)", "Status"]
    for i, (col, w) in enumerate(zip(rec_cols, [15, 22, 22, 16, 22, 22, 16, 16, 14]), 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
        _hcell(ws4, 1, i, col, F["HDR"], BD)

    for ri, d in enumerate(all_data, 2):
        its      = d.get("items", [])
        sum_av   = round(sum(x["Assess Value (INR)"] for x in its), 2)
        sum_duty = round(sum(x["Total Duty (INR)"]   for x in its), 2)
        hdr_av   = _safe_float(d.get("TOT ASS VAL", "0"))
        hdr_duty = _safe_float(d.get("TOT AMOUNT",  "0"))
        fine     = _safe_float(d.get("FINE",        "0"))
        diff_av   = round(sum_av   - hdr_av,   2)
        diff_duty = round(sum_duty - hdr_duty, 2)
        ok   = abs(diff_av) < 1 and abs(diff_duty) < 1
        fill = F["OK"] if ok else F["ERR"]
        for ci, v in enumerate(
            [d.get("BE No",""), sum_av, hdr_av, diff_av,
             sum_duty, hdr_duty, diff_duty, fine,
             "✔ MATCH" if ok else "⚠ MISMATCH"], 1
        ):
            fmt = "#,##0.00" if 2 <= ci <= 8 else None
            al  = "right"   if 2 <= ci <= 8 else "center"
            _dcell(ws4, ri, ci, v, BD, fill=fill, fmt=fmt, align=al, bold=(ci == 9))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    st.title("🛃 Indian Customs – Bill of Entry PDF Extractor")
    st.caption(
        "Upload one or more Warehouse BE (Out-of-Charge) PDFs. "
        "The app extracts header, invoice & item-level data and produces a "
        "formatted Excel report with duty reconciliation."
    )

    with st.expander("ℹ️ How to use & what gets extracted"):
        st.markdown("""
**Extracted sections**

| Section | Fields |
|---|---|
| **BE Header** | BE No/Date/Type, Port, IEC, GSTIN, CB, Importer, Country of Origin/Consignment, Port of Loading/Shipment, IGM, MAWB, GW, Packages, Exchange Rate, Duty totals, OOC details, Container & Seal |
| **Invoices** | Invoice No, Date, Amount, Currency (all invoices, any number) |
| **Items** | Inv S.No, Item S.No, CTH, Description, Qty (KGS), Unit Price, COO, **Assess Value (INR)**, **Total Duty (INR)** |
| **Duty Recon** | Sum of item-level Assess Values vs `18.TOT.ASS VAL`; Sum of item-level Duties vs `19.TOT.AMOUNT`. Difference should be **0** — if not, it should equal `17.FINE`. |

**Excel output sheets:** BE Headers · All Invoices · All Items · Duty Reconciliation
        """)

    uploaded_files = st.file_uploader(
        "Upload Bill of Entry PDF(s)",
        type="pdf",
        accept_multiple_files=True,
        help="You can select multiple PDFs at once.",
    )

    if not uploaded_files:
        st.info("👆 Upload one or more BE PDFs to get started.")
        return

    all_data: list = []
    progress = st.progress(0, text="Processing PDFs…")

    for idx, f in enumerate(uploaded_files):
        with st.spinner(f"Parsing {f.name} …"):
            try:
                d = extract_be_data(f.read())
                d["_filename"] = f.name
                all_data.append(d)
                st.success(
                    f"✅ **{f.name}** — BE {d.get('BE No', '?')}  |  "
                    f"{len(d.get('invoices', []))} invoice(s)  |  "
                    f"{len(d.get('items', []))} item(s) parsed"
                )
            except Exception as exc:
                st.error(f"❌ {f.name}: {exc}")
        progress.progress((idx + 1) / len(uploaded_files), text="Processing PDFs…")

    progress.empty()

    if not all_data:
        return

    st.divider()
    tab1, tab2, tab3, tab4 = st.tabs(
        ["📋 BE Header", "🧾 Invoices", "📦 Items", "⚖️ Duty Reconciliation"]
    )

    with tab1:
        rows = [{
            "BE No": d.get("BE No"), "BE Date": d.get("BE Date"),
            "Importer": d.get("Importer"), "CB Name": d.get("CB Name"),
            "IGM No": d.get("IGM No"), "IGM Date": d.get("IGM Date"),
            "INW Date": d.get("INW Date"),
            "MAWB No": d.get("MAWB No"), "MAWB Date": d.get("MAWB Date"),
            "GW (KGS)": d.get("GW (KGS)"), "Packages": d.get("Packages"),
            "Exchange Rate": d.get("Exchange Rate"),
            "Container No": d.get("Container No"), "Seal No": d.get("Seal No"),
            "TOT ASS VAL": d.get("TOT ASS VAL"),
            "TOT AMOUNT": d.get("TOT AMOUNT"), "FINE": d.get("FINE"),
            "OOC No": d.get("OOC No"), "OOC Date": d.get("OOC Date"),
        } for d in all_data]
        st.dataframe(pd.DataFrame(rows), use_container_width=True)

    with tab2:
        inv_rows = [{"BE No": d.get("BE No"), **inv}
                    for d in all_data for inv in d.get("invoices", [])]
        if inv_rows:
            df_inv = pd.DataFrame(inv_rows)
            st.dataframe(df_inv, use_container_width=True)
            st.metric("Total Invoice Value (USD)",
                      f"$ {df_inv['Invoice Amount (USD)'].sum():,.2f}")
        else:
            st.info("No invoice data could be parsed.")

    with tab3:
        item_rows = [{"BE No": d.get("BE No"), **it}
                     for d in all_data for it in d.get("items", [])]
        if item_rows:
            df_it = pd.DataFrame(item_rows)
            st.dataframe(df_it, use_container_width=True)
            c1, c2 = st.columns(2)
            c1.metric("Total Assess Value (INR)",
                      f"₹ {df_it['Assess Value (INR)'].sum():,.2f}")
            c2.metric("Total Duty (INR)",
                      f"₹ {df_it['Total Duty (INR)'].sum():,.2f}")
        else:
            st.info("No item-level duty data could be parsed. "
                    "The PDF must be text-based (not a scanned image).")

    with tab4:
        rec_rows = []
        for d in all_data:
            its      = d.get("items", [])
            sum_av   = round(sum(x["Assess Value (INR)"] for x in its), 2)
            sum_duty = round(sum(x["Total Duty (INR)"]   for x in its), 2)
            hdr_av   = _safe_float(d.get("TOT ASS VAL", "0"))
            hdr_duty = _safe_float(d.get("TOT AMOUNT",  "0"))
            fine     = _safe_float(d.get("FINE",        "0"))
            diff_av   = round(sum_av   - hdr_av,   2)
            diff_duty = round(sum_duty - hdr_duty, 2)
            ok = abs(diff_av) < 1 and abs(diff_duty) < 1
            rec_rows.append({
                "BE No": d.get("BE No"),
                "Sum AV – Items": sum_av,  "TOT ASS VAL – Header": hdr_av,  "Diff AV": diff_av,
                "Sum Duty – Items": sum_duty, "TOT AMOUNT – Header": hdr_duty, "Diff Duty": diff_duty,
                "FINE (Header)": fine,
                "Status": "✔ MATCH" if ok else "⚠ MISMATCH",
            })

        def _colour(row):
            c = "#E2EFDA" if row["Status"] == "✔ MATCH" else "#FFE0E0"
            return [f"background-color:{c}"] * len(row)

        st.dataframe(
            pd.DataFrame(rec_rows).style.apply(_colour, axis=1),
            use_container_width=True,
        )
        for row in rec_rows:
            if row["Status"] != "✔ MATCH":
                st.warning(
                    f"BE {row['BE No']}: Duty difference = ₹ {row['Diff Duty']:,.2f}. "
                    f"FINE in header = ₹ {row['FINE (Header)']:,.2f}. "
                    + ("✔ Matches FINE field."
                       if abs(abs(row["Diff Duty"]) - row["FINE (Header)"]) < 1
                       else "⚠ Does NOT match FINE – please review manually.")
                )

    st.divider()
    st.download_button(
        label="⬇️  Download Excel Report",
        data=build_excel(all_data),
        file_name="BE_Extract_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
